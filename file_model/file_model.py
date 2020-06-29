import os
import sys
import pandas as pd
from openpyxl import load_workbook


class FileModel:

    def __init__(self):
        self.destination_folder = str()
        self.list_files = list()
        self.sheets_in_wb = list()
        self.sheets_to_use = list()
        self.wb = str()
        self.wb_name = str()
        self.data_dict = dict()

    def get_folder(self):
        # ask for the folder and ensure input is a dir path
        self.formats(clear_window=True)
        self.formats(print_title="DIRECTORY PATH")
        destination_folder = input(r"Paste the destination folder path: ")
        while not "\\" in destination_folder:
            destination_folder = input("Must be a full directory path, C/users/you/folder/etc: ")
        # change cwd to that path
        try:
            os.chdir(destination_folder)
        except:
            destination_folder = input(r"ERROR: Not a directory path. Try again: ")
        # assign the destination folder in case needed again
        self.destination_folder = destination_folder

    def return_list_files(self):
        self.get_folder()
        for file in os.listdir():
            if 'xls' in file:
                self.list_files.append(file)

    def make_sheet_names(self):
        for sheet in self.wb.sheetnames:
            self.sheets_in_wb.append(sheet)

    def valid_chars_checker(self, inp):
        # check if valid chars passed
        valid_chars = [n for n in range(100)]
        valid_chars.insert(0, ',')
        if inp == 'all':
            return True
        else:
            for char in inp.split(','):
                try:
                    char = int(char)
                    if char not in valid_chars:
                        return False
                    if char > len(self.sheets_in_wb) - 1:
                        print("\n!!! You've input a number greater than what's available!!!\n")
                        return False
                except ValueError:
                    return False
        return True

    @staticmethod
    def valid_column_selection(df, selection):
        # check if valid chars passed
        valid_chars = [n for n in range(100)]
        valid_chars.insert(0, ',')
        if selection == 'all':
            return True
        else:
            for char in selection.split(','):
                try:
                    char = int(char)
                    if char not in valid_chars:
                        return False
                    if char > len(df.columns) - 1:
                        print("\n!!!You've input a number greater than the total number of columns available!!!\n")
                        return False
                except ValueError:
                    return False
        return True

    @staticmethod
    def formats(print_title=None, print_sub_header=None, print_hashes=False, print_dashes=False, clear_window=True):
        """
        Prints things for good format in cmd
        """
        form_hashes = "\n" + "#" * 100 + "\n"
        form_clear_window = "\n" * 100
        form_dashes = "\n" + "-" * 100 + "\n"

        if print_hashes:
            print(form_hashes)
        elif print_dashes:
            print(form_dashes)
        elif print_title is not None:
            print(form_hashes)
            print(f"{' ' * (60 - (len(print_title)))}{print_title.upper()}")
            print(form_hashes)
        elif print_sub_header is not None:
            print(form_dashes)
            print(f"{' ' * (60 - (len(print_sub_header)))}{print_sub_header.upper()}")
            print(form_dashes)
        elif clear_window:
            print(form_clear_window)

    def select_sheets_to_keep(self):
        # create the list of file
        self.return_list_files()
        # show the user what filenames were found
        # make clear space for visual purposes
        self.formats(print_title="LIST OF WORKBOOKS")
        print("Found the following workbooks.\n"
              "The script will automatically iterate over each one: you don't need to select a workbook.\n"
              "If you want to avoid certain workbooks then remove them from the directory & rerun this script.\n")
        for n in range(len(self.list_files)):
            print(f"\t{n}. {self.list_files[n]}")
        # iterate over all files
        for file in self.list_files:
            self.formats(print_title=f"FILE: {file}")
            self.wb_name = file
            # load the wb
            self.wb = load_workbook(file)
            # create the sheetnames to check if want to use
            self.make_sheet_names()
            # show user what file is being loaded
            print(f"Loading file: {file}\n")
            # if there's more than 1 sheet, ask user which they want to keep
            if len(self.sheets_in_wb) > 1:
                print("The following sheetnames were found\n")
                for sheet_num in range(len(self.sheets_in_wb)):
                    print(f"\t{sheet_num}. {self.sheets_in_wb[sheet_num]}")
                # begin process of asking which to keep
                print("\nWhich do you want to keep?\n"
                      "Type 'all' or comma separated numbers n,n,n")
                sheet_nums_to_keep = str(input("\ttype here: "))
                # if nothing typed then ask to type
                while len(sheet_nums_to_keep) < 1:
                    sheet_nums_to_keep = str(input("nothing typed, type again: "))

                # check valid chars have been passed and ask for re-entry if invalid
                if self.valid_chars_checker(sheet_nums_to_keep) is False:
                    breakout = False
                    while breakout is False:
                        sheet_nums_to_keep = str(input("Illegal character, type 'all' or n,n,n only: "))
                        if self.valid_chars_checker(sheet_nums_to_keep) is True:
                            breakout = True
                        else:
                            breakout = False

                # if reached this point then assumption is valid chars have been passed
                # checks will take place again at next inputs
                # now show the user what selection they made and offer chance for correction
                good_selection = False
                while good_selection is not True:
                    checker = str(input(f"\nYou selected | {sheet_nums_to_keep} | is this correct? y/n: "))
                    # make sure only y or n passed
                    while checker not in 'y Y n N'.split():
                        checker = str(input("just type 'y' or 'n': "))
                    # if no then give user chance to re input and check if valid each time
                    if checker in 'n N'.split():
                        # try/except to handle user error if presses 'enter' early
                        try:
                            # input
                            sheet_nums_to_keep = str(input("Ok, retype the numbers n,n,n etc or 'all': "))
                            # check if valid and demand re-entry if not
                            if self.valid_chars_checker(sheet_nums_to_keep) is False:
                                breakout = False
                                while breakout is False:
                                    sheet_nums_to_keep = str(input("Illegal character, type 'all' or n,n,n only: "))
                                    if self.valid_chars_checker(sheet_nums_to_keep) is True:
                                        breakout = True
                                    else:
                                        breakout = False
                        except:
                            # except clause to handle if user accidentally presses enter with no input
                            sheet_nums_to_keep = str(
                                input("You made a mistake? type comma separated numbers or 'all': "))
                        # back to start if error or invalid chars
                        good_selection = False
                    # checks if user good to proceed and appends sheetnames if so
                    # if made it this far then the input is valid

                    elif checker in 'y Y'.split():
                        good_selection = True
                        print(f"ok, keeping {sheet_nums_to_keep}")
                    if sheet_nums_to_keep == 'all':
                        for sheet in self.sheets_in_wb:
                            if sheet in self.sheets_to_use:
                                pass
                            else:
                                self.sheets_to_use.append(sheet)
                    else:
                        for num in sheet_nums_to_keep:
                            if num != 'all' and num != ',':
                                if self.sheets_in_wb[int(num)] not in self.sheets_to_use:
                                    self.sheets_to_use.append(self.sheets_in_wb[int(num)])
                                else:
                                    pass

            elif len(self.sheets_in_wb) == 1:
                pass
                print(f"Only one sheet found:\n\n\t 0. {self.sheets_in_wb[0]}")
                print("\nThis workbook has only one sheet."
                      "\nMoving to next file or next step if no more files.\n")
                self.sheets_to_use = self.sheets_in_wb
            # if made it this far then it means the sheets of the file have been selected
            # add to dict the filename and sheetnames for later use in pandas, clear other data
            # at that point the columnar criteria will be asked, not here
            self.data_dict[self.wb_name] = self.sheets_to_use
            self.data_dict[self.wb_name] = self.sheets_in_wb

            self.sheets_in_wb = []
            self.sheets_to_use = []

    def create_dataframes(self):
        # load in files and create data dict of workbook names and their respective sheets
        self.select_sheets_to_keep()

        # iterate over all keys (filenames) and load each respective worksheet as a dataframe
        for wb in self.data_dict.keys():
            for ws in self.data_dict[wb]:
                self.formats(print_title=f"WORKBOOK: {wb}\nWORKSHEET:{ws}")
                df = pd.read_excel(io=wb,
                                   sheet_name=ws,
                                   encoding=sys.getfilesystemencoding())
                self.formats(print_sub_header='COLUMN SELECTION')
                print("Here are a list of the column headers available\n"
                      "Input which you would like to keep.\n")
                list_columns = [c for c in df.columns]
                for n in range(len(list_columns)):
                    print(f"\t{n}. {list_columns[n]}")

                column_selection = str(input("\nSelect columns you want to keep by typing 'all' or n,n,n etc: "))
                print(f"\nYou have selected: {column_selection}\n")
                happy = str(input("Happy with that? y/n: "))
                while happy not in 'y Y n N'.split():
                    happy = str(input("Just type y or n: "))

                breakout = True
                if happy in 'y Y'.split():
                    # just proceed and check it's a valid choice
                    pass
                elif happy in 'n N'.split():
                    breakout = False
                    while breakout is False:
                        column_selection = str(input("\nOk, retype your selection: "))
                        print(f"You have now selected: {column_selection}\n")
                        happy = str(input("Are you happy with that? y/n: "))
                        if happy in 'n N'.split():
                            breakout = False
                        elif happy in 'y Y'.split():
                            print("ok, great")
                            breakout = True

                # check validity
                if self.valid_column_selection(df=df, selection=column_selection) is False:
                    breakout = False
                    while breakout is False:
                        column_selection = str(input("Illegal character, type 'all' or n,n,n only: "))
                        if self.valid_column_selection(df=df, selection=column_selection) is True:
                            breakout = True
                        else:
                            breakout = False
                # if made it this far then the column selection has taken place and now the df must be filtered
                # list of columns selected
                column_filter = []
                if column_selection == 'all':
                    print("\nOk, keeping all columns")
                    pass
                    # don't filter columns
                elif column_selection != 'all':
                    for num in column_selection.split(','):
                        column_filter.append(list_columns[int(num)])
                    print(f"\nOk, keeping columns {column_selection}")

                # filtering the dataframe according to the columns if necessary
                if column_selection != 'all':
                    df = df[column_filter]
                else:
                    pass
                self.formats(print_sub_header='INDEX COLUMN SELECTION')
                # list of columns in df
                list_columns = [c for c in df.columns]
                index_column = list_columns[0]
                print("\nYou need to select an index column.\n"
                      "This will be the primary grouping that will output in the text file.\n")
                for n in range(len(list_columns)):
                    print(f"\t{n}. {list_columns[n]}")
                try:
                    index_column = int(input("\nType the number of the column you want to choose: "))
                except:
                    index_column = int(input("Not a number! Type the COLUMN NUMBER you want to make index: "))
                while not isinstance(index_column, int):
                    index_column = int(input("Must be a whole number only: "))
                while index_column > len(list_columns) or index_column < 0:
                    index_column = int(input("You have chosen a column which doesn't exist. Try again: "))
                happy = str(input(f"\nAre you happy with column number {index_column}? y/n: "))
                breakout = True
                if happy in 'y Y'.split():
                    # proceed to set index column
                    pass
                elif happy in 'n N'.split():
                    breakout = False
                    while breakout is False:
                        index_column = str(input("\nOk, retype your selection: "))
                        print(f"You have now selected: {index_column}\n")
                        happy = str(input("Are you happy with that? y/n: "))
                        if happy in 'n N'.split():
                            breakout = False
                        elif happy in 'y Y'.split():
                            print("ok, great")
                            breakout = True
                # if made it this far then a valid integer selection has been made and the index will be set
                index_column = list_columns[index_column]
                df.set_index(index_column, drop=True, inplace=True)

                # proceed to create the text files
                txt_filename = wb.replace('xlsx', '').replace('.', '') + "_" + ws.replace('.', '') + ".txt"
                tcountry = ""
                vizaid = "-" * 10
                with open(txt_filename, 'w', encoding='utf-8') as txtfile:
                    for country in range(len(df.index)):
                        if tcountry == df.index[country]:
                            pass
                        else:
                            txtfile.write("#" * 20 + "\n")
                            tcountry = df.index[country]
                            txtfile.write(tcountry + "\n")
                            txtfile.write("#" * 20 + "\n\n")
                        for column in range(len(df.columns)):
                            txtfile.write(
                                f"{vizaid}{[c for c in df.columns][column]}: \n{df.iloc[country, column]}\n\n")

    def select_sheets_to_keep_fast(self):
        # create the list of file
        self.return_list_files()
        # show the user what filenames were found
        # make clear space for visual purposes
        self.formats(print_title="LIST OF WORKBOOKS")
        print("Found the following workbooks.\n"
              "The script will automatically iterate over each one: you don't need to select a workbook.\n"
              "If you want to avoid certain workbooks then remove them from the directory & rerun this script.\n")
        for n in range(len(self.list_files)):
            print(f"\t{n}. {self.list_files[n]}")
        # iterate over all files
        for file in self.list_files:
            self.formats(print_title=f"FILE: {file}")
            self.wb_name = file
            # load the wb
            self.wb = load_workbook(file)
            # create the sheetnames to check if want to use
            self.make_sheet_names()
            # show user what file is being loaded
            print(f"Loading file: {file}\n")
            # if there's more than 1 sheet, ask user which they want to keep
            # if made it this far then it means the sheets of the file have been selected
            # add to dict the filename and sheetnames for later use in pandas, clear other data
            # at that point the columnar criteria will be asked, not here
            self.data_dict[self.wb_name] = self.sheets_to_use
            self.data_dict[self.wb_name] = self.sheets_in_wb

            self.sheets_in_wb = []
            self.sheets_to_use = []

    def create_dataframes_fast(self):
        # load in files and create data dict of workbook names and their respective sheets
        self.select_sheets_to_keep_fast()

        # iterate over all keys (filenames) and load each respective worksheet as a dataframe
        for wb in self.data_dict.keys():
            for ws in self.data_dict[wb]:
                self.formats(print_title=f"WORKBOOK: {wb}\nWORKSHEET:{ws}")
                df = pd.read_excel(io=wb,
                                   sheet_name=ws,
                                   encoding=sys.getfilesystemencoding())
                # list of columns in df
                list_columns = [c for c in df.columns]
                # if made it this far then a valid integer selection has been made and the index will be set
                index_column = list_columns[0]
                df.set_index(index_column, drop=True, inplace=True)

                # proceed to create the text files
                txt_filename = wb.replace('xlsx', '').replace('.', '') + "_" + ws.replace('.', '') + ".txt"
                tcountry = ""
                vizaid = "-" * 10
                with open(txt_filename, 'w', encoding='utf-8') as txtfile:
                    for country in range(len(df.index)):
                        if tcountry == df.index[country]:
                            pass
                        else:
                            txtfile.write("#" * 20 + "\n")
                            tcountry = df.index[country]
                            txtfile.write(tcountry + "\n")
                            txtfile.write("#" * 20 + "\n\n")
                        for column in range(len(df.columns)):
                            txtfile.write(
                                f"{vizaid}{[c for c in df.columns][column]}: \n{df.iloc[country, column]}\n\n")